from ringcentral import SDK
import sys
import pandas as pd
from datetime import datetime, timedelta, timezone
from collections import defaultdict
from dotenv import load_dotenv
import os
from openpyxl import Workbook
from openpyxl.styles import Font
os.makedirs('rpt', exist_ok=True)

# Load variables from the .env file
load_dotenv()

# Retrieve the values
CLIENT_ID = os.getenv('CLIENT_ID')
CLIENT_SECRET = os.getenv('CLIENT_SECRET')
SERVER_URL = os.getenv('SERVER_URL')
JWT_TOKEN = os.getenv('JWT_TOKEN')

print(f"CLIENT_ID: {CLIENT_ID}")
print(f"SERVER_URL: {SERVER_URL}")
AGENT_EXTENSION = '11_digit_extension_id'  # Replace with the actual extension ID
if not CLIENT_ID or not CLIENT_SECRET or not SERVER_URL or not JWT_TOKEN:
    print("Error: Please set CLIENT_ID, CLIENT_SECRET, SERVER_URL, and JWT_TOKEN in the .env file.")
    sys.exit(1)

try:
    # Initialize SDK constructor
    rcsdk = SDK(CLIENT_ID, CLIENT_SECRET, SERVER_URL)
    platform = rcsdk.platform()

    # Authenticate
    platform.login(jwt=JWT_TOKEN)
    print("Authentication successful.")

    # Verify extension ID exists and get extension details 
    endpoint_extensions = '/restapi/v1.0/account/~/extension'
    response_extensions = platform.get(endpoint_extensions)
    raw_response = response_extensions.json()
    
    extensions = raw_response.records if hasattr(raw_response, 'records') else raw_response.get('records', []) if isinstance(raw_response, dict) else []

    extension_info = None
    for ext in extensions:
        ext_dict = ext.__dict__ if hasattr(ext, '__dict__') else ext
        ext_id = str(ext_dict.get('id') if isinstance(ext_dict, dict) else ext.id)
        if ext_id == AGENT_EXTENSION:
            extension_info = ext
            break

    if not extension_info:
        print(f"Error: Extension ID {AGENT_EXTENSION} not found in the account.")
        sys.exit(1)
    print(f"Extension {AGENT_EXTENSION} found.")

    # Get extension phone number or fallback to account main number 
    extension_phone = ''
    if hasattr(extension_info, 'contact') and hasattr(extension_info.contact, 'phoneNumber'):
        extension_phone = extension_info.contact.phoneNumber
    elif isinstance(extension_info, dict):
        extension_phone = extension_info.get('contact', {}).get('phoneNumber', '')

    if not extension_phone:
        account_endpoint = '/restapi/v1.0/account/~'
        account_response = platform.get(account_endpoint).json()
        extension_phone = getattr(account_response, 'mainNumber', 'Unknown') if hasattr(account_response, 'mainNumber') else account_response.get('mainNumber', 'Unknown') if isinstance(account_response, dict) else 'Unknown'
    print(f"Extension phone number: {extension_phone}")

    # Query message store for deleted messages
    endpoint = f'/restapi/v1.0/account/~/extension/{AGENT_EXTENSION}/message-store'
    params = {
        'perPage': 100,
        'page': 1,
        'dateFrom': (datetime.now(timezone.utc) - timedelta(days=30)).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'dateTo': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'availability': 'Deleted'  # Only fetch deleted messages
    }

    all_messages = []
    conversations = defaultdict(list)
    deleted_messages = []

    while True:
        response = platform.get(endpoint, params)
        messages_response = response.json()
        messages = messages_response.records if hasattr(messages_response, 'records') else messages_response.get('records', []) if isinstance(messages_response, dict) else []

        if not messages:
            break

        for msg in messages:
            msg_dict = msg.__dict__ if hasattr(msg, '__dict__') else msg
            message_status = msg_dict.get('messageStatus') if isinstance(msg_dict, dict) else msg.messageStatus
            message_type = msg_dict.get('type') if isinstance(msg_dict, dict) else msg.type
            from_obj = msg_dict.get('from', {}) if isinstance(msg_dict, dict) else msg.from_ or {}
            to_obj = msg_dict.get('to', [{}])[0] if isinstance(msg_dict, dict) else (msg.to[0] if msg.to else {})
            from_phone = from_obj.get('phoneNumber', '') if isinstance(from_obj, dict) else (from_obj.phoneNumber if hasattr(from_obj, 'phoneNumber') else '')
            to_phone = to_obj.get('phoneNumber', '') if isinstance(to_obj, dict) else (to_obj.phoneNumber if hasattr(to_obj, 'phoneNumber') else '')
            conversation_id = msg_dict.get('conversationId', '') if isinstance(msg_dict, dict) else getattr(msg, 'conversationId', '')

            # Use extension_phone if from_phone is empty
            if not from_phone:
                from_phone = extension_phone

            # Filter for SMS and Text messages
            if message_type not in ['SMS', 'Text']:
                continue

            # Determine the other party's phone number
            other_party = to_phone if msg_dict.get('direction', getattr(msg, 'direction', '')) == 'Outbound' else from_phone
            conversation_key = other_party

            message_data = {
                'MessageID': str(msg_dict.get('id', '') if isinstance(msg_dict, dict) else msg.id),
                'From': from_phone,
                'To': to_phone,
                'Text': msg_dict.get('subject', '') if isinstance(msg_dict, dict) else msg.subject,
                'Status': message_status,
                'Date': msg_dict.get('creationTime', '') if isinstance(msg_dict, dict) else msg.creationTime,
                'Type': message_type,
                'Availability': msg_dict.get('availability', '') if isinstance(msg_dict, dict) else msg.availability,
                'ConversationID': conversation_id,
                'Direction': msg_dict.get('direction', '') if isinstance(msg_dict, dict) else getattr(msg, 'direction', '')
            }
            all_messages.append(message_data)

            # Only add to conversations if deleted (redundant due to API filter)
            if message_data['Availability'] == 'Deleted':
                conversations[conversation_key].append(message_data)
                deleted_messages.append(message_data)

        paging = messages_response.paging if hasattr(messages_response, 'paging') else messages_response.get('paging', {}) if isinstance(messages_response, dict) else {}
        next_page_id = paging.get('nextPageId') if isinstance(paging, dict) else (paging.nextPageId if hasattr(paging, 'nextPageId') else None)
        if not next_page_id:
            break
        params['page'] = next_page_id or params['page'] + 1

    # Save all messages to Excel
    if all_messages:
        #print(f"Found {len(all_messages)} deleted messages.")
        df_all = pd.DataFrame(all_messages, columns=['MessageID', 'From', 'To', 'Text', 'Status', 'Date', 'Type', 'Availability', 'ConversationID', 'Direction'])
        df_all.to_excel('rpt/all_messages.xlsx', index=False, sheet_name='Deleted Messages')
        print("Excel file saved: all_messages.xlsx")
    else:
        print("No deleted messages found.")

    # Save deleted messages to Excel
    if deleted_messages:
        print(f"Found {len(deleted_messages)} deleted messages.")
        df_deleted = pd.DataFrame(deleted_messages, columns=['MessageID', 'From', 'To', 'Text', 'Status', 'Date', 'Type', 'Availability', 'ConversationID', 'Direction'])
        df_deleted.to_excel('rpt/deleted_messages.xlsx', index=False, sheet_name='Deleted Messages')
        print("Excel file saved: deleted_messages.xlsx")
    else:
        print("No deleted messages found.")

    # Save conversations to Excel
    conversation_dfs = []
    for contact, messages in conversations.items():
        sorted_messages = sorted(messages, key=lambda x: x['Date'])
        for msg in sorted_messages:
            msg['Conversation'] = f"{extension_phone} - {contact}"
            msg['Status'] = 'deleted'
        conversation_dfs.append(pd.DataFrame(sorted_messages, columns=['Conversation', 'MessageID', 'From', 'To', 'Text', 'Status', 'Date', 'Type', 'Availability', 'Direction']))
    
    if conversation_dfs:
        df_conversations = pd.concat(conversation_dfs, ignore_index=True)
        df_conversations.to_excel('rpt/conversations.xlsx', index=False, sheet_name='Conversations')
        print("Excel file saved: conversations.xlsx")
    else:
        print("No conversations found.")

    # Save readable conversations to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Readable Conversations"
    ws.append(['Message', 'Direction', 'Status'])  # Header row
    bold_font = Font(bold=True)

    row = 2  # Start after header
    for contact, messages in conversations.items():
        if not messages:
            continue
        ws[f'A{row}'] = f"Conversation with {contact}"
        ws[f'A{row}'].font = bold_font
        row += 1

        sorted_messages = sorted(messages, key=lambda x: x['Date'])
        for msg in sorted_messages:
            direction = 'sent' if msg['Direction'] == 'Outbound' else 'received'
            ws[f'A{row}'] = msg['Text']
            ws[f'B{row}'] = direction
            ws[f'C{row}'] = 'deleted'
            row += 1

        row += 1  # Empty row

    wb.save('readable.xlsx')
    print("Excel file saved: readable.xlsx")

    # Display conversations
    if conversations:
        for contact, messages in conversations.items():
            if not messages:
                continue
            print(f"\nConversation with {contact}:")
            sorted_messages = sorted(messages, key=lambda x: x['Date'])
            for msg in sorted_messages:
                direction = 'sent' if msg['Direction'] == 'Outbound' else 'received'
                print(f"{msg['Text']} - {direction}")
    else:
        print("No deleted messages found in conversations.")

except Exception as e:
    print(f"Error: {str(e)}")
    sys.exit(1)